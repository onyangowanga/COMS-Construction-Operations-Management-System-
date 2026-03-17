"""
Reporting Engine - Services

Report generation, export, caching, and scheduling services.
"""

import io
import os
import time
import hashlib
import pandas as pd
from decimal import Decimal
from datetime import datetime, timedelta
from django.db import transaction
from django.core.cache import cache
from django.utils import timezone
from django.conf import settings

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, PageBreak, Image
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from apps.reporting.models import Report, ReportSchedule, ReportExecution
from apps.common.services.code_generator import generate_report_code
from apps.reporting.selectors import (
    ProjectFinancialDataSelector,
    CashFlowDataSelector,
    VariationImpactDataSelector,
    SubcontractorPaymentDataSelector,
    DocumentAuditDataSelector,
    ProcurementDataSelector
)


class ReportService:
    """
    Main service for report generation and management.
    
    Handles report execution, caching, and file generation.
    """
    
    CACHE_PREFIX = 'report_'
    
    @staticmethod
    @transaction.atomic
    def create_report(
        organization,
        name,
        report_type,
        created_by,
        module='general',
        filters=None,
        columns=None,
        aggregations=None,
        group_by=None,
        code=None,
        description='',
        default_parameters=None,
        is_public=False,
        cache_duration=300
    ):
        """
        Create a new report configuration.
        
        Args:
            organization: Organization object
            name: Report name
            report_type: Report type from Report.ReportType
            created_by: User who creates the report
            description: Report description
            default_parameters: Default report parameters (dict)
            is_public: Whether report is visible to all users
            cache_duration: Cache duration in seconds
        
        Returns:
            Report: Created report object
        """
        generated_code, sequence, year = generate_report_code(organization)
        final_code = code or generated_code

        report = Report.objects.create(
            organization=organization,
            code=final_code,
            year=year,
            sequence=sequence,
            name=name,
            description=description,
            module=module,
            report_type=report_type,
            filters=filters or {},
            columns=columns or [],
            aggregations=aggregations or {},
            group_by=group_by or [],
            default_parameters=default_parameters or {},
            is_public=is_public,
            cache_duration=cache_duration,
            created_by=created_by
        )
        
        return report
    
    @staticmethod
    def generate_report(
        report,
        parameters,
        export_format,
        executed_by,
        use_cache=True
    ):
        """
        Generate a report with given parameters.
        
        Handles caching, data generation, and file export.
        
        Args:
            report: Report object
            parameters: Report parameters (dict)
            export_format: Export format (PDF, EXCEL, CSV, JSON)
            executed_by: User executing the report
            use_cache: Whether to use cached results
        
        Returns:
            ReportExecution: Execution record with file path
        """
        start_time = time.time()
        
        # Generate cache key
        cache_key = ReportService._generate_cache_key(
            report, parameters, export_format
        )
        
        # Check cache if enabled
        if use_cache and report.cache_duration > 0:
            cached_execution = cache.get(cache_key)
            if cached_execution:
                # Return cached execution
                return cached_execution
        
        # Create execution record
        execution = ReportExecution.objects.create(
            report=report,
            status=ReportExecution.Status.PROCESSING,
            export_format=export_format,
            parameters=parameters,
            cache_key=cache_key if use_cache else '',
            executed_by=executed_by
        )
        
        try:
            # Generate data based on report type
            data = ReportService._generate_report_data(
                report.report_type,
                parameters
            )
            
            # Export to requested format
            file_path, file_size = ReportService._export_report(
                report,
                data,
                export_format,
                execution
            )
            
            # Update execution record
            execution.status = ReportExecution.Status.COMPLETED
            execution.file_path = file_path
            execution.file_size = file_size
            execution.row_count = data.get('row_count', 0)
            execution.execution_time = time.time() - start_time
            execution.completed_at = timezone.now()
            execution.save()
            
            # Cache result if enabled
            if use_cache and report.cache_duration > 0:
                cache.set(cache_key, execution, report.cache_duration)
            
            return execution
            
        except Exception as e:
            # Mark as failed
            execution.status = ReportExecution.Status.FAILED
            execution.error_message = str(e)
            import traceback
            execution.stack_trace = traceback.format_exc()
            execution.completed_at = timezone.now()
            execution.save()
            raise
    
    @staticmethod
    def _generate_cache_key(report, parameters, export_format):
        """Generate unique cache key for report execution"""
        key_data = f"{report.id}:{export_format}:{str(parameters)}"
        hash_value = hashlib.md5(key_data.encode()).hexdigest()
        return f"{ReportService.CACHE_PREFIX}{hash_value}"
    
    @staticmethod
    def _generate_report_data(report_type, parameters):
        """
        Generate report data based on type.
        
        Routes to appropriate data selector.
        """
        project_id = parameters.get('project_id')
        organization_id = parameters.get('organization_id')
        start_date = parameters.get('start_date')
        end_date = parameters.get('end_date')
        
        # Parse dates if strings
        if isinstance(start_date, str):
            start_date = datetime.fromisoformat(start_date).date()
        if isinstance(end_date, str):
            end_date = datetime.fromisoformat(end_date).date()
        
        # Get project/organization objects
        from apps.projects.models import Project
        from apps.authentication.models import Organization
        
        project = Project.objects.get(id=project_id) if project_id else None
        organization = Organization.objects.get(id=organization_id) if organization_id else None
        
        # Route to appropriate generator
        generators = {
            Report.ReportType.PROJECT_FINANCIAL: lambda: ReportService._generate_project_financial(
                project, start_date, end_date
            ),
            Report.ReportType.CASH_FLOW: lambda: ReportService._generate_cash_flow(
                project, parameters.get('months', 12)
            ),
            Report.ReportType.VARIATION_IMPACT: lambda: ReportService._generate_variation_impact(
                project
            ),
            Report.ReportType.SUBCONTRACTOR_PAYMENT: lambda: ReportService._generate_subcontractor_payment(
                project, organization, start_date, end_date
            ),
            Report.ReportType.DOCUMENT_AUDIT: lambda: ReportService._generate_document_audit(
                project, organization, start_date, end_date
            ),
            Report.ReportType.PROCUREMENT_SUMMARY: lambda: ReportService._generate_procurement_summary(
                project, organization, start_date, end_date
            ),
        }
        
        generator = generators.get(report_type)
        if not generator:
            raise ValueError(f"Unknown report type: {report_type}")
        
        return generator()
    
    @staticmethod
    def _generate_project_financial(project, start_date, end_date):
        """Generate Project Financial Summary data"""
        summary = ProjectFinancialDataSelector.get_project_summary(
            project, start_date, end_date
        )
        
        return {
            'report_title': f'Project Financial Summary - {project.name}',
            'project': project,
            'summary': summary,
            'start_date': start_date,
            'end_date': end_date,
            'row_count': 1,
            'generated_at': timezone.now()
        }
    
    @staticmethod
    def _generate_cash_flow(project, months):
        """Generate Cash Flow Forecast data"""
        forecast = CashFlowDataSelector.get_cash_flow_forecast(project, months)
        
        return {
            'report_title': f'Cash Flow Forecast - {project.name}',
            'project': project,
            'forecast': forecast,
            'months': months,
            'row_count': len(forecast['revenue_forecast']) + len(forecast['expense_forecast']),
            'generated_at': timezone.now()
        }
    
    @staticmethod
    def _generate_variation_impact(project):
        """Generate Variation Impact Report data"""
        data = VariationImpactDataSelector.get_variation_summary(project)
        
        return {
            'report_title': f'Variation Impact Report - {project.name}',
            'project': project,
            'summary': data['summary'],
            'variations': data['variations'],
            'dataframe': data['dataframe'],
            'row_count': data['summary']['total_count'],
            'generated_at': timezone.now()
        }
    
    @staticmethod
    def _generate_subcontractor_payment(project, organization, start_date, end_date):
        """Generate Subcontractor Payment Report data"""
        data = SubcontractorPaymentDataSelector.get_subcontractor_payments(
            project, organization, start_date, end_date
        )
        
        return {
            'report_title': 'Subcontractor Payment Report',
            'project': project,
            'organization': organization,
            'claims': data['claims'],
            'summary': data['summary'],
            'dataframe': data['dataframe'],
            'start_date': start_date,
            'end_date': end_date,
            'row_count': len(data['summary']),
            'generated_at': timezone.now()
        }
    
    @staticmethod
    def _generate_document_audit(project, organization, start_date, end_date):
        """Generate Document Audit Report data"""
        data = DocumentAuditDataSelector.get_document_audit(
            project, organization, start_date, end_date
        )
        
        return {
            'report_title': 'Document Audit Report',
            'project': project,
            'organization': organization,
            'documents': data['documents'],
            'summary': data['summary'],
            'by_type': data['by_type'],
            'dataframe': data['dataframe'],
            'start_date': start_date,
            'end_date': end_date,
            'row_count': data['summary']['total_count'],
            'generated_at': timezone.now()
        }
    
    @staticmethod
    def _generate_procurement_summary(project, organization, start_date, end_date):
        """Generate Procurement Summary data"""
        data = ProcurementDataSelector.get_procurement_summary(
            project, organization, start_date, end_date
        )
        
        return {
            'report_title': 'Procurement Summary',
            'project': project,
            'organization': organization,
            'po_summary': data['po_summary'],
            'invoice_summary': data['invoice_summary'],
            'by_supplier': data['by_supplier'],
            'dataframe': data['dataframe'],
            'start_date': start_date,
            'end_date': end_date,
            'row_count': data['po_summary']['total_count'],
            'generated_at': timezone.now()
        }
    
    @staticmethod
    def _export_report(report, data, export_format, execution):
        """
        Export report data to specified format.
        
        Returns:
            tuple: (file_path, file_size)
        """
        if export_format == Report.ExportFormat.PDF:
            return PDFExportService.export(report, data, execution)
        elif export_format == Report.ExportFormat.EXCEL:
            return ExcelExportService.export(report, data, execution)
        elif export_format == Report.ExportFormat.CSV:
            return CSVExportService.export(report, data, execution)
        elif export_format == Report.ExportFormat.JSON:
            return JSONExportService.export(report, data, execution)
        else:
            raise ValueError(f"Unsupported export format: {export_format}")


class PDFExportService:
    """Service for exporting reports to PDF"""
    
    @staticmethod
    def export(report, data, execution):
        """
        Export report data to PDF.
        
        Returns:
            tuple: (file_path, file_size)
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is not installed. Install with: pip install reportlab")
        
        # Create reports directory if not exists
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        # Generate filename
        filename = f"report_{execution.id}.pdf"
        file_path = os.path.join(reports_dir, filename)
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(data['report_title'], title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Report info
        info_data = [
            ['Generated:', data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')],
            ['Project:', data.get('project').name if data.get('project') else 'All Projects'],
        ]
        if data.get('start_date'):
            info_data.append(['Period:', f"{data['start_date']} to {data['end_date']}"])
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#7f8c8d')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Add report-specific content
        if report.report_type == Report.ReportType.PROJECT_FINANCIAL:
            PDFExportService._add_financial_summary(story, data, styles)
        elif report.report_type == Report.ReportType.VARIATION_IMPACT:
            PDFExportService._add_variation_data(story, data, styles)
        elif report.report_type == Report.ReportType.SUBCONTRACTOR_PAYMENT:
            PDFExportService._add_subcontractor_data(story, data, styles)
        
        # Build PDF
        doc.build(story)
        
        # Save to file
        with open(file_path, 'wb') as f:
            f.write(buffer.getvalue())
        
        file_size = os.path.getsize(file_path)
        
        return (f'reports/{filename}', file_size)
    
    @staticmethod
    def _add_financial_summary(story, data, styles):
        """Add financial summary table to PDF"""
        summary = data['summary']
        
        # Summary table
        summary_data = [
            ['Financial Summary', ''],
            ['Contract Value', f"£{summary['contract_value']:,.2f}"],
            ['Approved Variations', f"£{summary['approved_variations']:,.2f}"],
            ['Revised Contract Value', f"£{summary['revised_contract_value']:,.2f}"],
            ['', ''],
            ['Total Certified', f"£{summary['total_certified']:,.2f}"],
            ['Total Received', f"£{summary['total_received']:,.2f}"],
            ['Outstanding Receivables', f"£{summary['outstanding_receivables']:,.2f}"],
            ['', ''],
            ['Total Costs', f"£{summary['total_costs']:,.2f}"],
            ['Total Paid Costs', f"£{summary['total_paid_costs']:,.2f}"],
            ['Outstanding Payables', f"£{summary['outstanding_payables']:,.2f}"],
            ['', ''],
            ['Gross Profit', f"£{summary['gross_profit']:,.2f}"],
            ['Profit Margin', f"{summary['profit_margin']:.2f}%"],
            ['Completion', f"{summary['completion_percentage']:.1f}%"],
        ]
        
        table = Table(summary_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0,0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecf0f1')]),
        ]))
        story.append(table)
    
    @staticmethod
    def _add_variation_data(story, data, styles):
        """Add variation data to PDF"""
        if not data.get('dataframe').empty:
            df = data['dataframe']
            # Display top 20 variations
            table_data = [['Reference', 'Title', 'Status', 'Amount']]
            for _, row in df.head(20).iterrows():
                table_data.append([
                    row.get('reference', ''),
                    row.get('title', '')[:50],
                    row.get('status', ''),
                    f"£{row.get('approved_amount' , row.get('estimated_amount', 0)):,.2f}"
                ])
            
            table = Table(table_data, colWidths=[1.5*inch, 3*inch, 1*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(table)
    
    @staticmethod
    def _add_subcontractor_data(story, data, styles):
        """Add subcontractor payment data to PDF"""
        if data.get('summary'):
            table_data = [['Subcontractor', 'Claimed', 'Certified', 'Paid', 'Outstanding']]
            for item in data['summary'][:20]:
                table_data.append([
                    item['subcontractor_name'][:30],
                    f"£{item['total_claimed']:,.0f}",
                    f"£{item['total_certified']:,.0f}",
                    f"£{item['total_paid']:,.0f}",
                    f"£{item['outstanding']:,.0f}"
                ])
            
            table = Table(table_data, colWidths=[2.5*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch])
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(table)


class ExcelExportService:
    """Service for exporting reports to Excel"""
    
    @staticmethod
    def export(report, data, execution):
        """
        Export report data to Excel.
        
        Returns:
            tuple: (file_path, file_size)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Install with: pip install openpyxl")
        
        # Create reports directory
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        # Generate filename
        filename = f"report_{execution.id}.xlsx"
        file_path = os.path.join(reports_dir, filename)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Styling
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = data['report_title']
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        # Report info
        ws['A3'] = 'Generated:'
        ws['B3'] = data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')
        
        if data.get('project'):
            ws['A4'] = 'Project:'
            ws['B4'] = data['project'].name
        
        # Add data based on report type
        row = 6
        if report.report_type == Report.ReportType.PROJECT_FINANCIAL:
            row = ExcelExportService._add_financial_data(ws, data, row, header_font, header_fill, border)
        elif report.report_type == Report.ReportType.VARIATION_IMPACT:
            row = ExcelExportService._add_variation_data_excel(ws, data, row, header_font, header_fill, border)
        elif report.report_type == Report.ReportType.SUBCONTRACTOR_PAYMENT:
            row = ExcelExportService._add_subcontractor_data_excel(ws, data, row, header_font, header_fill, border)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(file_path)
        file_size = os.path.getsize(file_path)
        
        return (f'reports/{filename}', file_size)
    
    @staticmethod
    def _add_financial_data(ws, data, start_row, header_font, header_fill, border):
        """Add financial summary to Excel"""
        summary = data['summary']
        
        # Headers
        ws.cell(row=start_row, column=1, value='Metric')
        ws.cell(row=start_row, column=2, value='Amount')
        ws.cell(row=start_row, column=1).font = header_font
        ws.cell(row=start_row, column=2).font = header_font
        ws.cell(row=start_row, column=1).fill = header_fill
        ws.cell(row=start_row, column=2).fill = header_fill
        
        # Data
        row = start_row + 1
        metrics = [
            ('Contract Value', summary['contract_value']),
            ('Approved Variations', summary['approved_variations']),
            ('Revised Contract Value', summary['revised_contract_value']),
            ('Total Certified', summary['total_certified']),
            ('Total Received', summary['total_received']),
            ('Outstanding Receivables', summary['outstanding_receivables']),
            ('Total Costs', summary['total_costs']),
            ('Gross Profit', summary['gross_profit']),
            ('Profit Margin', f"{summary['profit_margin']:.2f}%"),
        ]
        
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric)
            if isinstance(value, str):
                ws.cell(row=row, column=2, value=value)
            else:
                ws.cell(row=row, column=2, value=float(value))
                ws.cell(row=row, column=2).number_format = '£#,##0.00'
            row += 1
        
        return row + 2
    
    @staticmethod
    def _add_variation_data_excel(ws, data, start_row, header_font, header_fill, border):
        """Add variation data to Excel"""
        if not data.get('dataframe').empty:
            df = data['dataframe']
            
            # Headers
            headers = ['Reference', 'Title', 'Status', 'Estimated', 'Approved', 'Date']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Data
            row = start_row + 1
            for _, data_row in df.iterrows():
                ws.cell(row=row, column=1, value=data_row.get('reference', ''))
                ws.cell(row=row, column=2, value=data_row.get('title', ''))
                ws.cell(row=row, column=3, value=data_row.get('status', ''))
                ws.cell(row=row, column=4, value=float(data_row.get('estimated_amount', 0)))
                ws.cell(row=row, column=4).number_format = '£#,##0.00'
                ws.cell(row=row, column=5, value=float(data_row.get('approved_amount', 0) or 0))
                ws.cell(row=row, column=5).number_format = '£#,##0.00'
                ws.cell(row=row, column=6, value=str(data_row.get('created_at', '')))
                row += 1
            
            return row + 2
        
        return start_row
    
    @staticmethod
    def _add_subcontractor_data_excel(ws, data, start_row, header_font, header_fill, border):
        """Add subcontractor data to Excel"""
        if data.get('summary'):
            # Headers
            headers = ['Subcontractor', 'Claimed', 'Certified', 'Paid', 'Outstanding']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Data
            row = start_row + 1
            for item in data['summary']:
                ws.cell(row=row, column=1, value=item['subcontractor_name'])
                ws.cell(row=row, column=2, value=float(item['total_claimed']))
                ws.cell(row=row, column=2).number_format = '£#,##0.00'
                ws.cell(row=row, column=3, value=float(item['total_certified']))
                ws.cell(row=row, column=3).number_format = '£#,##0.00'
                ws.cell(row=row, column=4, value=float(item['total_paid']))
                ws.cell(row=row, column=4).number_format = '£#,##0.00'
                ws.cell(row=row, column=5, value=float(item['outstanding']))
                ws.cell(row=row, column=5).number_format = '£#,##0.00'
                row += 1
            
            return row + 2
        
        return start_row


class CSVExportService:
    """Service for exporting reports to CSV"""
    
    @staticmethod
    def export(report, data, execution):
        """Export to CSV using Pandas"""
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        filename = f"report_{execution.id}.csv"
        file_path = os.path.join(reports_dir, filename)
        
        # Get dataframe from data
        df = data.get('dataframe')
        if df is None or df.empty:
            # Create simple summary DataFrame
            df = pd.DataFrame([data.get('summary', {})])
        
        df.to_csv(file_path, index=False)
        file_size = os.path.getsize(file_path)
        
        return (f'reports/{filename}', file_size)


class JSONExportService:
    """Service for exporting reports to JSON"""
    
    @staticmethod
    def export(report, data, execution):
        """Export to JSON"""
        import json
        from decimal import Decimal
        
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        filename = f"report_{execution.id}.json"
        file_path = os.path.join(reports_dir, filename)
        
        # Custom JSON encoder for Decimal
        class DecimalEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                if isinstance(obj, datetime):
                    return obj.isoformat()
                return super().default(obj)
        
        # Prepare data (remove querysets)
        export_data = {
            'report_title': data.get('report_title'),
            'project': data.get('project').name if data.get('project') else None,
            'summary': data.get('summary', {}),
            'generated_at': data.get('generated_at').isoformat() if data.get('generated_at') else None
        }
        
        with open(file_path, 'w') as f:
            json.dump(export_data, f, cls=DecimalEncoder, indent=2)
        
        file_size = os.path.getsize(file_path)
        
        return (f'reports/{filename}', file_size)


class ReportScheduleService:
    """Service for managing scheduled reports"""
    
    @staticmethod
    @transaction.atomic
    def create_schedule(
        report,
        name,
        frequency,
        export_format,
        created_by,
        cron_expression='',
        parameters=None,
        delivery_method='EMAIL',
        recipients=None
    ):
        """Create a new report schedule"""
        schedule = ReportSchedule.objects.create(
            report=report,
            name=name,
            frequency=frequency,
            cron_expression=cron_expression,
            export_format=export_format,
            parameters=parameters or {},
            delivery_method=delivery_method,
            recipients=recipients or [],
            created_by=created_by
        )
        
        # Calculate next run
        schedule.next_run = ReportScheduleService._calculate_next_run(schedule)
        schedule.save()
        
        return schedule
    
    @staticmethod
    def _calculate_next_run(schedule):
        """Calculate next execution time for schedule"""
        now = timezone.now()
        
        if schedule.frequency == ReportSchedule.Frequency.DAILY:
            return now + timedelta(days=1)
        elif schedule.frequency == ReportSchedule.Frequency.WEEKLY:
            return now + timedelta(weeks=1)
        elif schedule.frequency == ReportSchedule.Frequency.MONTHLY:
            return now + timedelta(days=30)
        elif schedule.frequency == ReportSchedule.Frequency.QUARTERLY:
            return now + timedelta(days=90)
        elif schedule.frequency == ReportSchedule.Frequency.CUSTOM:
            # Parse cron expression (would need croniter library)
            # For now, default to 1 day
            return now + timedelta(days=1)
        
        return now + timedelta(days=1)
    
    @staticmethod
    def execute_due_schedules():
        """
        Execute all due report schedules.
        
        This should be called by a Celery periodic task.
        """
        due_schedules = ReportSchedule.objects.filter(
            is_active=True,
            next_run__lte=timezone.now()
        ).select_related('report')
        
        for schedule in due_schedules:
            try:
                # Execute report
                execution = ReportService.generate_report(
                    report=schedule.report,
                    parameters=schedule.parameters,
                    export_format=schedule.export_format,
                    executed_by=schedule.created_by,
                    use_cache=False  # Don't use cache for scheduled reports
                )
                
                # Update schedule
                schedule.last_run = timezone.now()
                schedule.next_run = ReportScheduleService._calculate_next_run(schedule)
                schedule.save()
                
                # Deliver report (email, storage, etc.)
                if schedule.delivery_method in [ReportSchedule.DeliveryMethod.EMAIL, ReportSchedule.DeliveryMethod.ALL]:
                    ReportScheduleService._deliver_by_email(execution, schedule)
                
            except Exception as e:
                # Log error (would integrate with logging system)
                print(f"Error executing schedule {schedule.id}: {e}")
                continue
    
    @staticmethod
    def _deliver_by_email(execution, schedule):
        """
        Deliver report via email.
        
        Placeholder for email delivery integration.
        """
        # Would integrate with Django email system
        # from django.core.mail import EmailMessage
        # email = EmailMessage(
        #     subject=f'Scheduled Report: {schedule.report.name}',
        #     body=f'Please find attached the report.',
        #     to=schedule.recipients,
        #     attachments=[(execution.file_path, ...)]
        # )
        # email.send()
        pass
